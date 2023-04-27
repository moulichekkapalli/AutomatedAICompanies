import numpy as np
import pandas as pd
import os
from app.database import fetchURLs, bulkstore
from app.encoded import encodeandindex
from app.crawler import classifycompany
from datetime import datetime, timezone
from app import app
import xlsxwriter
from openpyxl import load_workbook
import smtplib
from email.message import EmailMessage

app.config['MAIL_SERVER'] = os.environ.get("MAIL_SERVER")
app.config['MAIL_PORT'] = os.environ.get("MAIL_PORT")
app.config['MAIL_USERNAME'] = os.environ.get("MAIL_USERNAME")
app.config['MAIL_PASSWORD'] = os.environ.get("MAIL_PASSWORD")
app.config['MAIL_USE_TLS'] = os.environ.get("MAIL_USE_TLS")
app.config['MAIL_USE_SSL'] = os.environ.get("MAIL_USE_SSL")
msg = EmailMessage(app)


def fetchCompanies(query, top_k, index, model):
    query_vector = model.encode([query])
    top_k = index.search(query_vector, top_k)
    top_k_ids = top_k[1].tolist()[0]
    top_k_ids = list(np.unique(top_k_ids))
    print("top_k_ids len: ", len(top_k_ids))
    companies = [fetch_company_info(idx) for idx in top_k_ids]
    return companies


def fetch_company_info(dataframe_idx):
    # read pickle file as dataframe
    df = pd.read_pickle('app/others/companiesdata.pkl')
    info = df.iloc[dataframe_idx]
    comp_dict = {}
    comp_dict['url'] = info['url']
    comp_dict['summary'] = info['summary']
    comp_dict['category'] = info['category']
    comp_dict['name'] = info['title'].capitalize()
    return comp_dict


def validateCredentials(uname, pwd):
    return uname == os.environ.get("ADMIN_USERNAME") and pwd == os.environ.get("ADMIN_PASSWORD")


def validexcel(filename):
    excel_extensions = {'xls', 'xlsx', 'xlsm', 'xlsb', 'odf', 'ods', 'odt'}
    if not "." in filename:
        return False
    extension = filename.rsplit(".", 1)[1]
    if extension.lower() in excel_extensions:
        return True
    else:
        return False


def validexcelsheetname(file):
    # open an Excel file and return a workbook
    wb = load_workbook(file, read_only=True)
    if 'Sheet1' in wb.sheetnames:
        print('Sheet1 exists')
        return True
    else:
        return False


def prepareexcel(storedUrls, failedUrls):
    try:
        workbook = xlsxwriter.Workbook('app/others/exceldata.xlsx')
        worksheet = workbook.add_worksheet("Sheet1")
        worksheet.write('A1', 'URL')
        worksheet.write('B1', 'Remarks')
        row = 1
        col = 0

        # Iterate over the data and write it out row by row.
        for title, url, data, pcclass, meta, date in (storedUrls):
            worksheet.write(row, col, url)
            worksheet.write(row, col + 1, pcclass)
            row += 1

        for url, data in (failedUrls):
            worksheet.write(row, col, url)
            worksheet.write(row, col + 1, data)
            row += 1

        workbook.close()
    except Exception as error:
        print("Error while creating a excel file: ", error)
        raise Exception(error)


def send_mail_with_excel(urlsStoredSuccessfully):
    try:
        with app.app_context():
            msg = EmailMessage()
            msg['Subject'] = "Automated Identification of AI companies-bulk upload"
            msg['From'] = app.config['MAIL_USERNAME']
            msg['To'] = os.environ.get("MAIL_TO")

            if urlsStoredSuccessfully:
                msg.set_content(
                    "Dear Sir/Madam, \nGreetings from Automated Identification of AI companies \nYour input file was processed successfully. Kindly find the excel attachment below with remarks. \nThankyou for using Automated Identification of AI Companies. \nRegards, AI4SE Team ")
                with open('app/others/exceldata.xlsx', 'rb') as f:
                    file_data = f.read()
                msg.add_attachment(file_data, maintype="application",
                                   subtype="xlsx", filename='exceldata.xlsx')
            else:
                msg.set_content(
                    "Dear Sir/Madam, \nGreetings from Automated Identification of AI companies \nYour input file was not processed successfully. Please contact AI4SE team. \nThankyou for using Automated Identification of AI Companies. \nRegards, AI4SE Team ")

            with smtplib.SMTP_SSL(app.config['MAIL_SERVER'], app.config['MAIL_PORT']) as smtp:
                smtp.login(app.config['MAIL_USERNAME'],
                           app.config['MAIL_PASSWORD'])
                smtp.send_message(msg)

    except Exception as error:
        print("Error while sending mail: ", error)
        raise Exception(error)


def saveUrls(url_list):
    try:
        print("Executing background task of adding urls to db...")
        existing_urls = fetchURLs()
        companiesToBeAdded = []
        companiesFailedWhileCrawling = []
        for url in url_list:
            print("Company : ", url)
            if url not in existing_urls:
                print("Company not in database. Now crawling...")
                modelPred = classifycompany(str(url))
                if modelPred["status"] == 1:
                    print("Company crawled successfully...")
                    dt = datetime.now(timezone.utc)
                    companiesToBeAdded.append((modelPred["title"], modelPred["url"], modelPred["data"],
                                               modelPred["PCClass"], modelPred["summary"], dt))
                else:
                    print("Company crawling failed--->", modelPred["data"])
                    companiesFailedWhileCrawling.append(
                        (url, modelPred["data"]))
        print("Companies to be added are = ", len(companiesToBeAdded))
        if companiesToBeAdded:
            bulkstore(companiesToBeAdded)
            print("Urls saved to db. Now generating new index...")
            encodeandindex()
            print("New index is generated...")
            prepareexcel(companiesToBeAdded, companiesFailedWhileCrawling)
            print("Excel file is generated...")
            send_mail_with_excel(True)
            print("Excel file is sent via mail...")
    except Exception as error:
        print("Error in /saveUrls API:", error)
        send_mail_with_excel(False)
        raise Exception(error)
